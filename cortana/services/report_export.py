"""
Report Export Service - Generate PDF and Excel reports
"""
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib.units import inch
from reportlab.lib.enums import TA_CENTER, TA_RIGHT
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.chart import BarChart, PieChart, Reference
from datetime import datetime
from typing import Dict, List
import tempfile
import os
import logging

logger = logging.getLogger(__name__)


class ReportExporter:
    """Export financial reports to PDF and Excel formats"""

    def __init__(self):
        self.styles = getSampleStyleSheet()
        self.title_style = ParagraphStyle(
            'CustomTitle',
            parent=self.styles['Heading1'],
            fontSize=24,
            textColor=colors.HexColor('#1a73e8'),
            spaceAfter=30,
            alignment=TA_CENTER
        )
        self.heading_style = ParagraphStyle(
            'CustomHeading',
            parent=self.styles['Heading2'],
            fontSize=14,
            textColor=colors.HexColor('#1a73e8'),
            spaceAfter=12
        )

    def generate_pdf_report(self, summary_data: Dict, user_name: str, period: str = "weekly") -> str:
        """
        Generate a PDF report with financial summary

        Args:
            summary_data: Dictionary with financial data
            user_name: User's name
            period: "weekly" or "monthly"

        Returns:
            Path to the generated PDF file
        """
        # Create temp file
        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.pdf')
        pdf_path = temp_file.name
        temp_file.close()

        # Create PDF
        doc = SimpleDocTemplate(pdf_path, pagesize=letter)
        story = []

        # Title
        title = Paragraph(f"Cortana Financial Report", self.title_style)
        story.append(title)

        # Subtitle
        subtitle_text = f"{period.capitalize()} Report for {user_name}<br/>{datetime.now().strftime('%B %d, %Y')}"
        subtitle = Paragraph(subtitle_text, self.styles['Normal'])
        story.append(subtitle)
        story.append(Spacer(1, 0.3 * inch))

        # Summary Section
        story.append(Paragraph("Summary", self.heading_style))

        summary_data_table = [
            ["Total Income", f"${summary_data.get('total_income', 0):.2f}"],
            ["Total Expenses", f"${summary_data.get('total_expenses', 0):.2f}"],
            ["Net Balance", f"${summary_data.get('net_balance', 0):.2f}"]
        ]

        table = Table(summary_data_table, colWidths=[3 * inch, 2 * inch])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#f8f9fa')),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 12),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
            ('TOPPADDING', (0, 0), (-1, -1), 12),
            ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#dee2e6'))
        ]))
        story.append(table)
        story.append(Spacer(1, 0.3 * inch))

        # Expenses by Category
        expenses_by_category = summary_data.get('expenses_by_category', {})
        if expenses_by_category:
            story.append(Paragraph("Expenses by Category", self.heading_style))

            category_data = [["Category", "Amount", "Percentage"]]
            total_expenses = summary_data.get('total_expenses', 0)

            for category, amount in sorted(expenses_by_category.items(), key=lambda x: x[1], reverse=True):
                percentage = (amount / total_expenses * 100) if total_expenses > 0 else 0
                category_data.append([
                    category.capitalize(),
                    f"${amount:.2f}",
                    f"{percentage:.1f}%"
                ])

            category_table = Table(category_data, colWidths=[2.5 * inch, 1.5 * inch, 1.5 * inch])
            category_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1a73e8')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('ALIGN', (1, 0), (-1, -1), 'RIGHT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 0), (-1, -1), 11),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 10),
                ('TOPPADDING', (0, 0), (-1, -1), 10),
                ('BACKGROUND', (0, 1), (-1, -1), colors.white),
                ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#dee2e6'))
            ]))
            story.append(category_table)
            story.append(Spacer(1, 0.3 * inch))

        # Recent Transactions
        recent_transactions = summary_data.get('recent_transactions', [])
        if recent_transactions:
            story.append(Paragraph("Recent Transactions", self.heading_style))

            transaction_data = [["Date", "Description", "Category", "Amount"]]

            for trans in recent_transactions[:10]:  # Show last 10
                trans_date = trans.get('date', datetime.now())
                if hasattr(trans_date, 'strftime'):
                    date_str = trans_date.strftime('%b %d')
                else:
                    date_str = str(trans_date)

                transaction_data.append([
                    date_str,
                    trans.get('description', '')[:30],
                    trans.get('category', '').capitalize(),
                    f"${trans.get('amount', 0):.2f}"
                ])

            trans_table = Table(transaction_data, colWidths=[1 * inch, 2.5 * inch, 1.5 * inch, 1 * inch])
            trans_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1a73e8')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('ALIGN', (3, 0), (3, -1), 'RIGHT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 0), (-1, -1), 9),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
                ('TOPPADDING', (0, 0), (-1, -1), 8),
                ('BACKGROUND', (0, 1), (-1, -1), colors.white),
                ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#dee2e6')),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8f9fa')])
            ]))
            story.append(trans_table)

        # Footer
        story.append(Spacer(1, 0.5 * inch))
        footer_text = f"Generated by Cortana Finance Assistant<br/>{datetime.now().strftime('%B %d, %Y at %I:%M %p')}"
        footer = Paragraph(footer_text, self.styles['Normal'])
        story.append(footer)

        # Build PDF
        doc.build(story)
        logger.info(f"Generated PDF report: {pdf_path}")

        return pdf_path

    def generate_excel_report(self, summary_data: Dict, user_name: str, period: str = "weekly") -> str:
        """
        Generate an Excel report with financial summary and charts

        Args:
            summary_data: Dictionary with financial data
            user_name: User's name
            period: "weekly" or "monthly"

        Returns:
            Path to the generated Excel file
        """
        # Create temp file
        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
        excel_path = temp_file.name
        temp_file.close()

        # Create workbook
        wb = openpyxl.Workbook()

        # Summary Sheet
        ws_summary = wb.active
        ws_summary.title = "Summary"

        # Header
        ws_summary['A1'] = "Cortana Financial Report"
        ws_summary['A1'].font = Font(size=20, bold=True, color="1a73e8")
        ws_summary['A2'] = f"{period.capitalize()} Report for {user_name}"
        ws_summary['A2'].font = Font(size=12, italic=True)
        ws_summary['A3'] = datetime.now().strftime('%B %d, %Y')

        # Summary Data
        ws_summary['A5'] = "Summary"
        ws_summary['A5'].font = Font(size=14, bold=True, color="1a73e8")

        summary_rows = [
            ["Total Income", summary_data.get('total_income', 0)],
            ["Total Expenses", summary_data.get('total_expenses', 0)],
            ["Net Balance", summary_data.get('net_balance', 0)]
        ]

        for idx, (label, value) in enumerate(summary_rows, start=6):
            ws_summary[f'A{idx}'] = label
            ws_summary[f'B{idx}'] = value
            ws_summary[f'B{idx}'].number_format = '"$"#,##0.00'
            ws_summary[f'A{idx}'].font = Font(bold=True)

        # Expenses by Category
        expenses_by_category = summary_data.get('expenses_by_category', {})
        if expenses_by_category:
            ws_summary['A10'] = "Expenses by Category"
            ws_summary['A10'].font = Font(size=14, bold=True, color="1a73e8")

            # Headers
            ws_summary['A11'] = "Category"
            ws_summary['B11'] = "Amount"
            ws_summary['C11'] = "Percentage"
            for cell in ['A11', 'B11', 'C11']:
                ws_summary[cell].font = Font(bold=True)
                ws_summary[cell].fill = PatternFill(start_color="1a73e8", end_color="1a73e8", fill_type="solid")
                ws_summary[cell].font = Font(bold=True, color="FFFFFF")

            total_expenses = summary_data.get('total_expenses', 0)
            row_idx = 12

            for category, amount in sorted(expenses_by_category.items(), key=lambda x: x[1], reverse=True):
                percentage = (amount / total_expenses * 100) if total_expenses > 0 else 0
                ws_summary[f'A{row_idx}'] = category.capitalize()
                ws_summary[f'B{row_idx}'] = amount
                ws_summary[f'B{row_idx}'].number_format = '"$"#,##0.00'
                ws_summary[f'C{row_idx}'] = percentage / 100
                ws_summary[f'C{row_idx}'].number_format = '0.0%'
                row_idx += 1

        # Transactions Sheet
        recent_transactions = summary_data.get('recent_transactions', [])
        if recent_transactions:
            ws_trans = wb.create_sheet(title="Transactions")

            # Headers
            headers = ["Date", "Description", "Category", "Amount", "Type"]
            for col_idx, header in enumerate(headers, start=1):
                cell = ws_trans.cell(row=1, column=col_idx)
                cell.value = header
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="1a73e8", end_color="1a73e8", fill_type="solid")
                cell.alignment = Alignment(horizontal='center')

            # Data
            for row_idx, trans in enumerate(recent_transactions, start=2):
                trans_date = trans.get('date', datetime.now())
                if hasattr(trans_date, 'strftime'):
                    date_str = trans_date.strftime('%Y-%m-%d')
                else:
                    date_str = str(trans_date)

                ws_trans[f'A{row_idx}'] = date_str
                ws_trans[f'B{row_idx}'] = trans.get('description', '')
                ws_trans[f'C{row_idx}'] = trans.get('category', '').capitalize()
                ws_trans[f'D{row_idx}'] = trans.get('amount', 0)
                ws_trans[f'D{row_idx}'].number_format = '"$"#,##0.00'
                ws_trans[f'E{row_idx}'] = trans.get('type', 'expense').capitalize()

            # Auto-fit columns
            for column in ws_trans.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws_trans.column_dimensions[column_letter].width = adjusted_width

        # Auto-fit columns in summary sheet
        for column in ws_summary.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws_summary.column_dimensions[column_letter].width = adjusted_width

        # Save workbook
        wb.save(excel_path)
        logger.info(f"Generated Excel report: {excel_path}")

        return excel_path

    def cleanup_file(self, file_path: str):
        """Delete temporary file"""
        try:
            if os.path.exists(file_path):
                os.unlink(file_path)
                logger.info(f"Cleaned up file: {file_path}")
        except Exception as e:
            logger.error(f"Failed to cleanup file {file_path}: {str(e)}")
